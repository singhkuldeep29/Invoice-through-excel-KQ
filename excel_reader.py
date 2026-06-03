import openpyxl

# Canonical column names (what the Excel sheet should contain).
# Payor/Payer variants are both accepted — see _normalize().
EXPECTED_COLUMNS = [
    'S No.', 'Invoice No.', 'Invoice Date', 'Period', 'Expense Nature',
    'TDS Section', 'TDS Rate %', 'Payor Name', 'Payor Type', 'Payor PAN',
    'Payor Address', 'Payee Name', 'Payee Address', 'Payee PAN',
    'Payee Bank Name', 'Payee Bank A/c No.', 'Payee IFSC',
    'Gross Amount', 'TDS Amount', 'Net Payable', 'Generate?',
]

REQUIRED_FOR_GENERATION = [
    'Invoice No.', 'Invoice Date', 'Payee Name', 'Payor Name',
    'Payee PAN', 'Payor PAN', 'Gross Amount',
    'Payee Bank Name', 'Payee Bank A/c No.', 'Payee IFSC',
]


def _normalize(s):
    """Lowercase + strip for fuzzy column matching."""
    return str(s or '').strip().lower()


def _build_col_map(headers):
    """
    Returns {canonical_name: column_index} mapping.
    Accepts 'Payer Address' as alias for 'Payor Address'.
    Also accepts 'Description' column if present.
    """
    aliases = {
        'payer pan': 'Payor PAN',
        'payor pan': 'Payor PAN',
        'payer name': 'Payor Name',
        'payer type': 'Payor Type',
        'payer address': 'Payor Address',
        'payor address': 'Payor Address',
        'payee bank a/c no.': 'Payee Bank A/c No.',
        'payee bank ac no.': 'Payee Bank A/c No.',
        'payee bank account no.': 'Payee Bank A/c No.',
        's no.': 'S No.',
        's.no.': 'S No.',
        'tds rate %': 'TDS Rate %',
        'tds rate': 'TDS Rate %',
        'generate?': 'Generate?',
    }

    col_map = {}
    for idx, header in enumerate(headers):
        key = _normalize(header)
        # direct match against expected canonical names
        canonical = next((c for c in EXPECTED_COLUMNS if _normalize(c) == key), None)
        if canonical:
            col_map[canonical] = idx
            continue
        # alias match
        if key in aliases:
            col_map[aliases[key]] = idx
            continue
        # Description column (optional, user-added)
        if key == 'description':
            col_map['Description'] = idx

    return col_map


def read_invoices(xlsx_path):
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = _build_col_map(headers)

    missing = [c for c in EXPECTED_COLUMNS if c not in col_map]
    if missing:
        raise ValueError(f"Missing columns in Excel: {', '.join(missing)}")

    has_description = 'Description' in col_map

    def get(row, col_name):
        idx = col_map.get(col_name)
        if idx is None:
            return None
        val = row[idx]
        return val

    invoices = []
    warnings = []
    skipped_n = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue  # blank row

        generate = _normalize(get(row, 'Generate?'))
        if generate != 'y':
            skipped_n += 1
            continue

        data = {col: get(row, col) for col in EXPECTED_COLUMNS}
        data['Description'] = get(row, 'Description') if has_description else None

        missing_fields = [f for f in REQUIRED_FOR_GENERATION if not data.get(f)]
        if missing_fields:
            warnings.append(f"Row {row_num}: skipped — missing: {', '.join(missing_fields)}")
            continue

        invoices.append(data)

    return invoices, warnings, skipped_n
